# -*- coding: utf-8 -*-
import os
from tkinter import *
from tkinter import ttk
from tkinter.ttk import *
from tkinter.ttk import Entry
from PIL import Image, ImageTk
# adding pdf library for the report
from reportlab.pdfgen import canvas
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.platypus import Image as img
from reportlab.lib.units import inch
# from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import datetime as dt
import requests
from io import BytesIO
# import os
# from openpyxl import Workbook
import openpyxl

################################################

# excel_file = "C:\\python\\NADI\\grammh_paragwghs v2.0\\pinakas metrisewn springs 1 21.10.2022.xlsx"
# wb = openpyxl.load_workbook(excel_file, read_only=True)
# ws = wb.active
#
# for row in ws.iter_rows(1):
#     for cell in row:
#         if cell.value == "2400x2200":
#             print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value,
#                   row[7].value, row[8].value, row[9].value, row[10].value, row[11].value)

###############################################


# ΤΟ FILEPATHTEST ΠΑΙΡΝΕΙ ΤΟ ΧΡΗΣΤΗ ΑΥΤΟΜΑΤΑ. ΘΑ ΤΟ ΧΡΗΣΙΜΟΠΟΙΗΣΩ ΓΙΑ ΝΑ ΒΡΙΣΚΕΙ ΤΟ ΠΡΟΓΡΑΜΜΑ ΤΙΣ ΕΙΚΟΝΕΣ ΚΑΙ ΤΑ
# ΑΡΧΕΙΑ ΠΟΥ ΧΡΕΙΑΖΕΤΑΙ ΟΤΑΝ ΕΙΝΑΙ ΝΑ ΤΟ ΕΓΚΑΤΑΣΤΗΣΩ ΣΕ ΑΛΛΑ PC. ΠΧ ΘΑ ΤΑ ΒΑΛΩ ΟΛΑ ΜΕΣΑ ΣΕ ΚΑΠΟΙΟ ΦΑΚΕΛΟ ΣΕ PROGRAM
# FILES (ΣΑΝ ΤΑ ΑΡΧΕΙΑ ΕΓΚΑΤΑΣΤΑΣΗΣ ΤΩΝ ΚΑΝΟΝΙΚΩΝ ΠΡΟΓΡΑΜΜΑΤΩΝ
# user = os.getlogin()
# print(user)
# filepath = 'C:\\Program Files\\nadi_program'
# filepathtest = 'C:\\'+user+'\\Desktop'
# print(filepathtest)

root = Tk()
logo = PhotoImage(file='C:\\python\\NADI\\grammh_paragwghs v2.0\\logo.png')
root.iconphoto(False, logo)
# root.geometry("750x250")
root.resizable(True, True)
my_notebook = ttk.Notebook(root)
my_notebook.pack(expand=1, fill=BOTH)

# create tabs
client_tab = ttk.Frame(my_notebook)
my_notebook.add(client_tab, text="ΣΤΟΙΧΕΙΑ ΠΕΛΑΤΗ")
general_tab = ttk.Frame(my_notebook)
my_notebook.add(general_tab, text="ΓΕΝΙΚΑ")

root.title('Υπολογισμός Υλικών Κατασκευής Πόρτας')
var = StringVar()

imagelogo = (Image.open('C:\\python\\NADI\\grammh_paragwghs v2.0\\logo.png'))
resized_imagelogo = imagelogo.resize((160, 80), Image.ANTIALIAS)
new_resized_imagelogo = ImageTk.PhotoImage(resized_imagelogo)

# add paragraph pdf style
my_Style = ParagraphStyle('My Para Style',
                          fontName='Times-Roman',
                          fontSize=12,
                          aligment=0,
                          borderWidth=0,
                          borderColor='#FFFF00',
                          backColor='#F1F1F1',
                          borderPadding=(0, 0, 0),
                          leading=20
                          )
my_Style2 = ParagraphStyle('My Para Style',
                           fontName='Times-Roman',
                           fontSize=12,
                           aligment=0,
                           borderWidth=0,
                           borderColor='#FFFF00',
                           backColor='#F1F1F1',
                           borderPadding=(0, 0, 0),
                           leading=20
                           )
my_Style3 = ParagraphStyle('My Para Style',
                           fontName='Times-Roman',
                           fontSize=13,
                           aligment=0,
                           borderWidth=0,
                           borderColor='#FFFFFF',
                           backColor='#FFFFFF',
                           borderPadding=(0, 0, 0),
                           leading=20
                           )
file_path = "C:\\python\\NADI\\grammh_paragwghs v2.0\\"  # path of the generated file
width, height = A4  # size of the file


# Δημιουργία PDF
def gen_pdf():
    banner_file = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\nadi_banner.jpg'
    door_photo = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\door_photo.png'
    # can = canvas.Canvas(packet)
    # can.drawImage(banner_file, 0, 0, width=120, preserveAspectRatio=True, mask='auto')

    # SEARCHING IF A FOLDER EXISTS. IF IT DOES NOT IT CREATES IT SO THAT WE WILL SORT EVERY
    # DIFFERENT CLIENT'S PDF INTO A FOLDER WITH HIS NAME.IF THE FOLDER ALREADY EXIST IT SKIPS THE CREATION OF THE FOLDER
    folder_name = name_value.get()
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    #
    c = canvas.Canvas(file_path + folder_name + '\\' + name_value.get() + '_' + f"{date:%d-%m-%Y}" + '.pdf',
                      pagesize=A4)
    x_start = 0
    y_start = 0
    info = '<br/> &nbsp &nbsp &nbsp Ημερομηνια: ' + date_value.get() + '&nbsp &nbsp &nbsp Τηλεφωνο: ' \
           + phone_value.get() + '&nbsp &nbsp &nbsp Σταθερο: ' + home_phone_value.get() \
           + '<br/> &nbsp &nbsp &nbsp Οδος: ' + address_value.get() + '&nbsp &nbsp &nbsp Πολη: ' + city_value.get() \
           + '&nbsp &nbsp &nbsp Τ.Κ.: ' + zip_code_value.get() + '&nbsp &nbsp &nbsp Χωρα: ' + country_value.get() \
           + '<br/> &nbsp &nbsp &nbsp Πελατης: ' + name_value.get() + '&nbsp &nbsp &nbsp Α.Φ.Μ.: ' + afm_value.get() \
           + '&nbsp &nbsp &nbsp ΔΟΥ: ' + doy_value.get() + '&nbsp &nbsp &nbsp Εταιρεια: ' + company_value.get()

    p1 = Paragraph(info, my_Style2)
    p1.wrapOn(c, 595, 50)
    p1.drawOn(c, width - 595, height - 180)
    # ΤΑ ΕΚΤΥΠΩΝΕΙ ΚΑΝΟΝΙΚΑ ΚΑΙ ΟΧΙ ΣΕ ΠΙΝΑΚΑ ΟΠΩΣ ΤΟ ΑΠΟ ΚΑΤΩ ΤΟΥ
    # dimensions = 'Διαστασεις: ' + width_value.get() + 'Πx' + height_value.get() +'Υ <br/>' + 'Αριθμος Πανελ(500mm):'+\
    #              panel_number_value.get() + '<br/> Αριθμος Πανελ(610mm):' + panel_number_value2.get()
    # p2 = Paragraph(dimensions, my_Style)
    # p2.wrapOn(c, 595, 50)  # dimension of the paragraph
    # p2.drawOn(c, width - 595, height - 240)  # location of paragraph

    # image of the pdf, banner topside
    c.drawImage(banner_file, x_start, y_start, width=595, height=1565, preserveAspectRatio=True, mask='auto')
    # DOOR IMAGE 1ST PAGE END
    # door_photo_resized = door_photo.resize(100, 50, Image.ANTIALIAS)
    c.drawImage(door_photo, x_start + 208, y_start, width=200, height=200, preserveAspectRatio=True, mask='auto')
    if lift_system_types_combobox.current() == 0:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\sl.png'
        c.drawImage(lift_type_img, x_start + 350, y_start + 390, width=200, height=200, preserveAspectRatio=True,
                    mask='auto')
    elif lift_system_types_combobox.current() == 1:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\lh.png'
        c.drawImage(lift_type_img, x_start + 350, y_start + 390, width=200, height=200, preserveAspectRatio=True,
                    mask='auto')
    elif lift_system_types_combobox.current() == 2:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\vl.png'
        c.drawImage(lift_type_img, x_start + 320, y_start + 420, width=200, height=150, preserveAspectRatio=True,
                    mask='auto')
    elif lift_system_types_combobox.current() == 3:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\hl.png'
        c.drawImage(lift_type_img, x_start + 320, y_start + 420, width=200, height=150, preserveAspectRatio=True,
                    mask='auto')
    elif lift_system_types_combobox.current() == 4:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\hlp.png'
        c.drawImage(lift_type_img, x_start + 320, y_start + 420, width=200, height=150, preserveAspectRatio=True,
                    mask='auto')
    elif lift_system_types_combobox.current() == 5:
        lift_type_img = 'C:\\python\\NADI\\grammh_paragwghs v2.0\\slp.png'
        c.drawImage(lift_type_img, x_start + 320, y_start + 420, width=200, height=150, preserveAspectRatio=True,
                    mask='auto')

    # string for image value width
    width_to_photo = width_value.get() + 'mm'
    width_to_photo_paragraph = Paragraph(width_to_photo, my_Style3)
    width_to_photo_paragraph.wrapOn(c, 80, 50)
    width_to_photo_paragraph.drawOn(c, width - 305, height - 800)
    # string 90 degree rotation for image value height
    c.saveState()
    c.rotate(90)
    c.drawRightString(width - 520, height - 1253, str(height_value.get()) + 'mm')
    c.restoreState()

    # ΑΡΧΙΚΟΣ ΠΙΝΑΚΑΣ ΠΛΗΡΟΦΟΡΙΩΝ
    c.drawCentredString(x_start + 160, y_start + 640, 'ΓΕΝΙΚΕΣ ΠΛΗΡΟΦΟΡΙΕΣ')
    data = [
        ['Καθαρο Ανοιγμα (mm)', width_value.get() + 'Πx' + height_value.get() + 'Y'],
        ['Κωδικος Χρωματος', color_value.get()],
        ['Πανελ 50mm', panel_number_value.get()],
        ['Πανελ 61mm', panel_number_value2.get()],
        ['Τυπος Πανελ', panel_types_combobox.get()],
        ['Βαρος Πανελ/m2 (kg)', panel_weight_value.get()],
        ['Συνολικo Βαρος Πανελ (kg)', panel_weight_total_value.get()],
        ['Τυπος Πορτας', lift_system_types_combobox.get()],
        ['Λειτουργια', operation_combobox.get()],
        ['Τυπος Μοτερ', moter_model_combobox.get()],
        # ['Θεση Ελατηριων', track_system_types_combobox.get()],
        ['Ενισχυση Κατω Πανελ', bottom_reinforcement_combobox.get()]
    ]

    table = Table(data)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.darkgrey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        # ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        # ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black)

        # ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        # ('BOX', (0, 0), (-1, -1), 0.25, colors.black)
    ])
    table.setStyle(table_style)
    table.wrapOn(c, 500, 150)
    table.drawOn(c, width - 585, height - 404)
    # ΤΕΛΟΣ ΑΡΧΙΚΟΥ ΠΙΝΑΚΑ ΠΛΗΡΟΦΟΡΙΩΝ
    #     c.showPage()  # Close the current page and  start on a new page.

    # 6/2/2023 ΔΙΑΘΕΣΙΜΟΙ ΚΩΔΙΚΟΙ ΟΔΗΓΩΝ ΟΙΚΙΑΣ
    # ΟΡΙΖΟΝΤΙΟΙ ΟΔΗΓΟΙ
    horizontal_track_code = ''
    if int(height_value.get()) <= 2000:
        horizontal_track_code = 'RSCHM10'
    elif 2000 < int(height_value.get()) <= 2125:
        horizontal_track_code = 'RSCHM20'
    elif 2125 < int(height_value.get()) <= 2250:
        horizontal_track_code = 'RSCHM30'
    elif 2250 < int(height_value.get()) <= 2375:
        horizontal_track_code = 'RSCHM40'
    elif 2375 < int(height_value.get()) <= 2500:
        horizontal_track_code = 'RSCHM50'
    elif 2500 < int(height_value.get()) <= 2750:
        horizontal_track_code = 'RSCHM60'
    elif 2750 < int(height_value.get()) <= 3000:
        horizontal_track_code = 'RSCHM70'

    horizontal_track_dimensions = 0
    if horizontal_track_code == 'RSCHM10':
        horizontal_track_dimensions = 2000
    elif horizontal_track_code == 'RSCHM20':
        horizontal_track_dimensions = 2125
    elif horizontal_track_code == 'RSCHM30':
        horizontal_track_dimensions = 2250
    elif horizontal_track_code == 'RSCHM40':
        horizontal_track_dimensions = 2375
    elif horizontal_track_code == 'RSCHM50':
        horizontal_track_dimensions = 2500
    elif horizontal_track_code == 'RSCHM60':
        horizontal_track_dimensions = 2750
    elif horizontal_track_code == 'RSCHM70':
        horizontal_track_dimensions = 3000

    # ΚΑΘΕΤΟΙ ΟΔΗΓΟΙ
    vertical_track_code = ''
    if int(height_value.get()) <= 2000:
        vertical_track_code = 'RSCVBC10Z'
    elif 2000 < int(height_value.get()) <= 2125:
        vertical_track_code = 'RSCVBC20Z'
    elif 2125 < int(height_value.get()) <= 2250:
        vertical_track_code = 'RSCVBC30Z'
    elif 2250 < int(height_value.get()) <= 2375:
        vertical_track_code = 'RSCVBC40Z'
    elif 2375 < int(height_value.get()) <= 2500:
        vertical_track_code = 'RSCVBC50Z'
    elif 2500 < int(height_value.get()) <= 2750:
        vertical_track_code = 'RSCVBC60Z'
    elif 2750 < int(height_value.get()) <= 3000:
        vertical_track_code = 'RSCVBC70Z'

    vertical_track_dimensions = 0
    if vertical_track_code == 'RSCVBC10Z':
        vertical_track_dimensions = 2000
    elif vertical_track_code == 'RSCVBC20Z':
        vertical_track_dimensions = 2125
    elif vertical_track_code == 'RSCVBC30Z':
        vertical_track_dimensions = 2250
    elif vertical_track_code == 'RSCVBC40Z':
        vertical_track_dimensions = 2375
    elif vertical_track_code == 'RSCVBC50Z':
        vertical_track_dimensions = 2500
    elif vertical_track_code == 'RSCVBC60Z':
        vertical_track_dimensions = 2750
    elif vertical_track_code == 'RSCVBC70Z':
        vertical_track_dimensions = 3000

    # ΤΕΛΟΣ ΔΙΑΘΕΣΙΜΩΝ ΟΔΗΓΩΝ ΟΙΚΙΑΣ

    # ΑΡΧΗ ΠΙΝΑΚΑ ΚΟΠΩΝ ΟΔΗΓΩΝ
    # print(type(vertical_track_code_dimensions))
    int_height_value = int(height_value.get())
    c.drawCentredString(x_start + 460, y_start + 640, 'ΠΙΝΑΚΑΣ ΚΟΠΗΣ ΟΔΗΓΩΝ')
    kopes_odigwn_data = [
        ['Καθετοι:',
         vertical_track_code + '(1set) κοπη: ' + str(vertical_track_dimensions - int_height_value) + 'mm'],
        ['Οριζοντιοι:',
         horizontal_track_code + '(1set) κοπη: ' + str(horizontal_track_dimensions - int_height_value) + 'mm'],
        ['Λαστιχα οδηγων', '(1L&1R) ' + height_value.get() + 'mm'],

    ]

    kopes_odigwn_table = Table(kopes_odigwn_data)
    kopes_odigwn_table.setStyle(table_style)
    kopes_odigwn_table.wrapOn(c, 100, 150)
    kopes_odigwn_table.drawOn(c, width - 308, height - 260)
    # ΤΕΛΟΣ ΠΙΝΑΚΑ ΚΟΠΩΝ ΟΔΗΓΩΝ
    # koph_panw_panel_main = on_field_change_height(koph_panw_panel)
    # ΠΙΝΑΚΑΣ ΚΟΠΩΝ ΠΑΝΕΛ
    c.drawCentredString(x_start + 250, y_start + 395, 'ΠΙΝΑΚΑΣ ΚΟΠΗΣ ΠΑΝΕΛ')
    kopes_data = [
        ['Κοπη Πλατους', str(int(width_value.get()) + 40) + 'mm'],
        ['Κοπη Υψους', 'Φυρα:' + str((int(panel_number_value.get()) * 500 + int(panel_number_value2.get()) * 610)
                                     - int(height_value.get())) + 'mm'],
        ['Τελειωματα',
         'Αρ. 50mm:' + str(round(float(panel_number_value.get()))) + ' Δεξ. 50mm:' +
         str(round(float(panel_number_value.get()))) + ' Αρ. 61mm:' + str(round(float(panel_number_value2.get())))
         + ' Δεξ. 61mm:' + str(round(float(panel_number_value2.get())))],
        ['Αλουμηνια', 'Πανω:' + str(int(width_value.get())) + 'mm Κατω:' + str(int(width_value.get())) +
         'mm'],
        ['Κοπη Λαστιχα', 'Πανω:' + str(int(width_value.get())) + 'mm Κατω:' + str(int(width_value.get())) + 'mm'],
        ['Dection', dection_types_combobox.get()]
    ]
    kopes_table = Table(kopes_data)
    kopes_table.setStyle(table_style)
    kopes_table.wrapOn(c, 200, 150)
    kopes_table.drawOn(c, width - 585, height - 560)
    # ΤΕΛΟΣ ΠΙΝΑΚΑ ΚΟΠΩΝ
    c.showPage()  # Close the current page and  start on a new page.

    # ΑΡΧΗ ΠΙΝΑΚΑ ΑΞΟΝΑ
    c.drawString(x_start + 43, y_start + 790, 'ΠΙΝΑΚΑΣ ΑΞΟΝΑ')
    shaft_table_data = [
        ['Συρματοσχοινα', lifting_cable_types_combobox.get()],
        ['Συνδεσμος Αξονα', shaft_coupler_combobox.get()],
        ['Τυπος Αξονα', shaft_type_values_combobox.get()]
    ]

    shaft_table_data = Table(shaft_table_data)
    shaft_table_data.setStyle(table_style)
    shaft_table_data.wrapOn(c, 200, 150)
    shaft_table_data.drawOn(c, width - 585, height - 120)
    # ΤΕΛΟΣ ΠΙΝΑΚΑ ΑΞΟΝΑ

    # SEARCHING IN EXCEL THE GIVEN VALUE OF WIDTH x HEIGHT AND PRINTS THE RESULTS OF THE ROWS THAT MUCH IT Αρχικά
    # δηλώνουμε που βρίσκετε το αρχείο μας (excel_file) και μετά το ανοίγουμε επισημαίνοντας οτι είναι read only (wb)
    # με το ws δηλώνουμε πιο spreadsheet του excel μας θέλουμε και δημιουργούμε μια κενή λίστα την οποία αργότερα θα
    # γεμίσουμε με τις τιμές που θα προκύψουν από την αναζήτηση στο excel. Μέσα στη for row αναζητάμε σε κάθε γραμμή
    # του excel σε κάθε κελί αν η τιμή του κελιού είναι ίση με το πλάτος και το ύψος που έχει δώσει ο χρήστης τότε οι
    # τιμές των γραμμών που καταχωρούνται μέσα στη λίστα μας.
    excel_file = "C:\\python\\NADI\\grammh_paragwghs v2.0\\pinakas metrisewn springs 1 3.11.22.xlsx"
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb.active

    spring_cut_table_data = [
        ['Διαστασεις (ΠxΥ)', 'Αρ. Ελατ.', 'Θεση', 'Εσ. Διαμ.', 'Διαμ. Συρμ.', 'Μηκος', 'Στροφες',
         'Βαρος', 'Κυκλοι Ζωης'],
    ]
    # Μετατροπή του width_value.get() και του height_value.get() ανάλογα με το υπόλοιπο της διαίρεσης τους με το 100
    # έτσι ώστε αν έχουν υπόλοιπο μεγαλύτερο απο 40 να ανεβαίνει στο επόμενο νούμερο ο υπολογισμός των ελατηρίων.
    # Πχ.: αν έχουμε σαν δοθείσα τιμή width_value.get() = 2065 ο υπολογισμός των ελατηρίων να γίνεται για 2100
    # ενώ αν έχουμε 2035 ο υπολογισμός των ελατηρίων να γίνεται για 2000.
    # Το παρακάτω παράδειγμα αναφέρεται στη λειτουργία αυτή με ένα απλό user input.

    # user_input = input('Dose ari8mo:')
    # print('user_input:', user_input)
    #
    # ypoloipo = int(user_input) % 100
    # print('ypoloipo:', ypoloipo)
    # epomenh_katostada = 100 - ypoloipo
    # print('epomenh_katostada:', epomenh_katostada)

    # if epomenh_katostada < 40:
    #     user_input = int(user_input) + int(epomenh_katostada)
    #     print('if print1:', user_input)
    # else:
    #     user_input = int(user_input) - int(ypoloipo)
    #     print('if print2:', user_input)
    user_given_width_value = width_value.get()
    print('user_given_width_value:', user_given_width_value)
    ypoloipo_user_given_width_value = int(user_given_width_value) % 100
    print('ypoloipo_user_given_width_value:', ypoloipo_user_given_width_value)
    epomenh_katostada_user_given_width_value = 100 - ypoloipo_user_given_width_value
    print('epomenh_katostada_user_given_width_value:', epomenh_katostada_user_given_width_value)
    if epomenh_katostada_user_given_width_value <= 60:
        user_given_width_value = int(user_given_width_value) + int(epomenh_katostada_user_given_width_value)
        print('IF condition 1 for epomenh katostada <= 40:user_given_width_value=', user_given_width_value)
    elif 100 > epomenh_katostada_user_given_width_value > 60:
        user_given_width_value = int(user_given_width_value) - int(ypoloipo_user_given_width_value)
        print('IF condition 2 for epomenh katostada > 40:user_given_width_value=', user_given_width_value)
    elif epomenh_katostada_user_given_width_value == 100:
        print('IF condition 3 for epomenh katostada == 100:user_given_width_value=', user_given_width_value)
        user_given_width_value = width_value.get()
        pass
    user_given_height_value = height_value.get()
    print('user_given_height_value:', height_value.get())
    ypoloipo_user_given_height_value = int(user_given_height_value) % 100
    print('ypoloipo_user_given_height_value', ypoloipo_user_given_height_value)
    epomenh_katostada_user_given_height_value = 100 - ypoloipo_user_given_height_value
    print('epomenh_katostada_user_given_height_value:', epomenh_katostada_user_given_height_value)
    if epomenh_katostada_user_given_height_value <= 60:
        user_given_height_value = int(user_given_height_value) + int(epomenh_katostada_user_given_height_value)
        print('IF condition 1 for epomenh katostada <= 40:user_given_height_value=', user_given_height_value)
    elif 100 > epomenh_katostada_user_given_height_value > 60:
        user_given_height_value = int(user_given_height_value) - int(ypoloipo_user_given_height_value)
        print('IF condition 2 for epomenh katostada > 40:user_given_height_value=', user_given_height_value)
    elif epomenh_katostada_user_given_height_value == 100:
        print('IF condition 3 for epomenh katostada == 100:user_given_height_value=', user_given_height_value)
        user_given_height_value = height_value.get()
        pass

    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == str(user_given_width_value) + "x" + str(user_given_height_value):
                ##########################################
                # EDW MESA PERNIETAI TO PERIEXOMENO KA8E GRAMMHS MESA STH LISTA MAS KAI META EKTOS THS IF
                # APO KATW GINETAI H DHMIOURGIA TOY PINAKA
                spring_cut_table_data.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value,
                                              row[5].value, row[7].value, row[8].value, row[9].value])

    #
    table_style_springs = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.darkgrey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        # ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        # ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black)

        # ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        # ('BOX', (0, 0), (-1, -1), 0.25, colors.black)
    ])
    #
    spring_cut_table_data_table = Table(spring_cut_table_data, repeatRows=1)
    spring_cut_table_data_table.setStyle(table_style_springs)
    print()
    print('mhkos listas epilogwn elathriwn apo excel (panta einai to pragmatiko len+1 epeidh h prwth grammh '
          'einai oi titloi)', len(spring_cut_table_data))
    print()

    if 18 <= len(spring_cut_table_data) <= 21:
        c.drawString(x_start + 150, y_start + 690, 'ΠΙΝΑΚΑΣ ΣΥΝΔΙΑΣΜΩΝ ΕΛΑΤΗΡΙΩΝ ΓΙΑ ΤΙΣ ΔΙΑΣΤΑΣΕΙΣ')
        spring_cut_table_data_table.wrapOn(c, 400, 150)  # poso xwro pianei sth selida
        spring_cut_table_data_table.drawOn(c, width - 590, height - 530)  # pou einai sth selida
    elif 17 >= len(spring_cut_table_data) >= 15:
        c.drawString(x_start + 150, y_start + 700, 'ΠΙΝΑΚΑΣ ΣΥΝΔΙΑΣΜΩΝ ΕΛΑΤΗΡΙΩΝ ΓΙΑ ΤΙΣ ΔΙΑΣΤΑΣΕΙΣ')
        spring_cut_table_data_table.wrapOn(c, 400, 150)  # poso xwro pianei sth selida
        spring_cut_table_data_table.drawOn(c, width - 590, height - 450)  # pou einai sth selida
    elif 14 >= len(spring_cut_table_data) >= 12:
        c.drawString(x_start + 150, y_start + 700, 'ΠΙΝΑΚΑΣ ΣΥΝΔΙΑΣΜΩΝ ΕΛΑΤΗΡΙΩΝ ΓΙΑ ΤΙΣ ΔΙΑΣΤΑΣΕΙΣ')
        spring_cut_table_data_table.wrapOn(c, 400, 150)  # poso xwro pianei sth selida
        spring_cut_table_data_table.drawOn(c, width - 590, height - 400)  # pou einai sth selida
    elif 11 >= len(spring_cut_table_data) >= 1:
        c.drawString(x_start + 150, y_start + 705, 'ΠΙΝΑΚΑΣ ΣΥΝΔΙΑΣΜΩΝ ΕΛΑΤΗΡΙΩΝ ΓΙΑ ΤΙΣ ΔΙΑΣΤΑΣΕΙΣ')
        spring_cut_table_data_table.wrapOn(c, 400, 150)  # poso xwro pianei sth selida
        spring_cut_table_data_table.drawOn(c, width - 590, height - 340)  # pou einai sth selida
    elif len(spring_cut_table_data) > 21:
        c.drawString(x_start + 150, y_start + 700, 'ΠΙΝΑΚΑΣ ΣΥΝΔΙΑΣΜΩΝ ΕΛΑΤΗΡΙΩΝ ΓΙΑ ΤΙΣ ΔΙΑΣΤΑΣΕΙΣ')
        spring_cut_table_data_table.wrapOn(c, 400, 150)  # poso xwro pianei sth selida
        spring_cut_table_data_table.drawOn(c, width - 590, height - 600)  # pou einai sth selida
    print()
    print('ARXH LISTAS', spring_cut_table_data_table, 'TELOS LISTAS')
    print()
    address = address_value.get() + ',' + city_value.get() + ',' + zip_code_value.get()
    # Get the static map image
    # print(address)
    if address != '':
        print('oxi kenh ', address)
        map_url = f"https://maps.googleapis.com/maps/api/staticmap?center={address}&zoom=16&size=650x300&maptype=road" \
                  f"map&markers=color:red%7Clabel:S%7C{address}&key=AIzaSyBNPA5EqlS8f9aMy0ZrzXeXCpH5eBcQbJk"

        # map_url = "https://maps.googleapis.com/maps/api/staticmap?center=Tositsa+5,Pefka+570+10&zoom=16&size=600x300&
        # maptype=roadmap&markers=color:blue%7Clabel:S%7C40.702147,-74.015794&markers=color:green%7Clabel:G%7C40.711614,-74.012318&markers=color:red%7Clabel:C%7C40.718217,-73.998284&key=AIzaSyBNPA5EqlS8f9aMy0ZrzXeXCpH5eBcQbJk"
        map_image = Image.open(BytesIO(requests.get(map_url).content))
        pdf_file = BytesIO()
        c.drawInlineImage(map_image, 0.5, 0.5, width=650, height=300, anchor='center')
    # address = address_value.get()
    # Get the static map image
    # print(address)
    # if address != '':
    #     print('oxi kenh ', address)
    #     map_url = f"https://maps.googleapis.com/maps/api/staticmap?center={address}&zoom=16&size=600x500&maptype=road
    #     map&markers=color:red%7Clabel:S%7C{address}&key=AIzaSyBNPA5EqlS8f9aMy0ZrzXeXCpH5eBcQbJk"
    #
    #     # map_url = "https://maps.googleapis.com/maps/api/staticmap?center=Tositsa+5,Pefka+570+10&zoom=16&size=600x300
    #     &maptype=roadmap&markers=color:blue%7Clabel:S%7C40.702147,-74.015794&markers=color:green%7Clabel:G%7C40.711614,-74.012318&markers=color:red%7Clabel:C%7C40.718217,-73.998284&key=AIzaSyBNPA5EqlS8f9aMy0ZrzXeXCpH5eBcQbJk"
    #     map_image = Image.open(BytesIO(requests.get(map_url).content))
    #     pdf_file = BytesIO()
    #     c.drawInlineImage(map_image, 0.5, 0.5, width=500, height=300,)

    ##########################################
    # END OF SEARCHING

    # 27/12/2022 ΒΑΣΙΚΑ ΥΛΙΚΑ ΠΙΝΑΚΑΣ ΜΕ ΕΙΚΟΝΕΣ
    c.showPage()

    # Importing images
    it_424_czn10_r_afb1 = img('424CZN10R__afb1.jpg')
    it_424hzn10r = img('424HZN10R__afb1.jpg')
    it_1085n_2510 = img('1085N-2510__afb1.jpg')
    it_215c_b6500 = img('215C-B6500__afb1.jpg')
    it_rscvb50z = img('RSCVB50Z__afb1.jpg')
    it_rschm50 = img('RSCHM50__afb1.jpg')
    it_870rs = img('870RS__afb1.jpg')
    it_574rn = img('574RN__afb1.jpg')
    it_574s = img('574S__afb1.jpg')
    it_577looprol = img('577__afb1.jpg')
    it_liftmaster = img('liftmaster_afb1.jpg')
    it_usa8r = img('USA8R__afb1.jpg')
    it_usa8l = img('USA8L__afb1.jpg')
    it_421k = img('421K__afb1.jpg')
    it_usa_b = img('USA-B__afb1.jpg')
    it_ff05nl12 = img('FF05NL12__afb1.jpg')
    it_k24_4800_ts = img('k24-4800ts__afb1.jpg')
    it_vl45_1_300gs = img('VL45-1-3000GS__afb1.jpg')
    it_1036 = img('1036__afb2.jpg')
    it_1035_dw = img('1035dw-35m_copy.jpg')
    it_a1f = img('a1ef.jpg')
    it_1085rbn2035 = img('1085RBN2035__afb1.png')
    it_702gbl750 = img('702GBL2750__afb1.jpg')
    it_703st = img('703ST__afb1.jpg')
    it_629v = img('629V__afb1.jpg')
    it_italdoor_panel = img('italdoorAPZDogaLarga.jpg')
    moter = img('c40.jpg')
    aluminium_profiles = img('1038-4040.jpg')

    # fixing image sizes
    it_424_czn10_r_afb1.drawHeight = 0.35 * inch
    it_424_czn10_r_afb1.drawWidth = 0.35 * inch

    it_424hzn10r.drawHeight = 0.35 * inch
    it_424hzn10r.drawWidth = 0.35 * inch

    it_1085n_2510.drawHeight = 0.35 * inch
    it_1085n_2510.drawWidth = 0.35 * inch

    it_215c_b6500.drawHeight = 0.35 * inch
    it_215c_b6500.drawWidth = 0.35 * inch

    it_rscvb50z.drawHeight = 0.35 * inch
    it_rscvb50z.drawWidth = 0.35 * inch

    it_rschm50.drawHeight = 0.35 * inch
    it_rschm50.drawWidth = 0.35 * inch

    it_870rs.drawHeight = 0.35 * inch
    it_870rs.drawWidth = 0.35 * inch

    it_574rn.drawWidth = 0.35 * inch
    it_574rn.drawHeight = 0.35 * inch

    it_574s.drawHeight = 0.35 * inch
    it_574s.drawWidth = 0.35 * inch

    it_577looprol.drawHeight = 0.35 * inch
    it_577looprol.drawWidth = 0.35 * inch

    it_liftmaster.drawHeight = 0.35 * inch
    it_liftmaster.drawWidth = 0.35 * inch

    it_usa8r.drawHeight = 0.35 * inch
    it_usa8r.drawWidth = 0.35 * inch

    it_usa8l.drawHeight = 0.35 * inch
    it_usa8l.drawWidth = 0.35 * inch

    it_421k.drawHeight = 0.35 * inch
    it_421k.drawWidth = 0.35 * inch

    it_usa_b.drawHeight = 0.35 * inch
    it_usa_b.drawWidth = 0.35 * inch

    it_ff05nl12.drawHeight = 0.35 * inch
    it_ff05nl12.drawWidth = 0.35 * inch

    it_k24_4800_ts.drawHeight = 0.35 * inch
    it_k24_4800_ts.drawWidth = 0.35 * inch

    it_vl45_1_300gs.drawHeight = 0.35 * inch
    it_vl45_1_300gs.drawWidth = 0.35 * inch

    it_1036.drawHeight = 0.35 * inch
    it_1036.drawWidth = 0.35 * inch

    it_1035_dw.drawHeight = 0.35 * inch
    it_1035_dw.drawWidth = 0.35 * inch

    it_a1f.drawHeight = 0.35 * inch
    it_a1f.drawWidth = 0.35 * inch

    it_1085rbn2035.drawHeight = 0.35 * inch
    it_1085rbn2035.drawWidth = 0.35 * inch

    it_702gbl750.drawHeight = 0.35 * inch
    it_702gbl750.drawWidth = 0.35 * inch

    it_703st.drawHeight = 0.35 * inch
    it_703st.drawWidth = 0.35 * inch

    it_629v.drawHeight = 0.35 * inch
    it_629v.drawWidth = 0.35 * inch

    it_italdoor_panel.drawHeight = 0.35 * inch
    it_italdoor_panel.drawWidth = 0.35 * inch

    moter.drawHeight = 0.35 * inch
    moter.drawWidth = 0.35 * inch

    aluminium_profiles.drawHeight = 0.35 * inch
    aluminium_profiles.drawWidth = 0.35 * inch

    table_style2 = TableStyle([
        # ('BACKGROUND', (0, 0), (0, -1), colors.darkgrey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        # ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        # ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black)

        # ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        # ('BOX', (0, 0), (-1, -1), 0.25, colors.black)
    ])

    moter_yes_or_no = ''
    if operation_combobox.current() == 1:
        moter_yes_or_no = '1 ' + str(moter_model_combobox.get())
    else:
        moter_yes_or_no = '-'
    # c.drawCentredString(x_start + 300, y_start + 805, 'ΠΙΝΑΚΑΣ ΕΞΑΡΤΗΜΑΤΩΝ')

    # checks if the selected all variable is set to TRUE so that it PRINTS ALL ITEMS
    # ΚΑΝΟΝΙΖΕΙ ΠΟΙΑ ΑΠΟ ΤΑ ΥΛΙΚΑ ΘΑ ΕΚΤΥΠΩΘΟΥΝ ΑΝΑΛΟΓΑ ΜΕ ΤΟ ΤΙ ΕΧΕΙ ΕΠΙΛΕΓΕΙ ΣΤΑ CHECKBOXES
    items_data = [['']]
    checkbox_all_data = checkbox_var_all.get()
    checkbox_shaft_data = checkbox_var_shaft.get()
    checkbox_panel_data = checkbox_var_panel.get()
    checkbox_rail_data = checkbox_var_rail.get()
    checkbox_general_data = checkbox_var_general.get()

    if checkbox_all_data == 1:
        print('checkbox_all variable is TRUE')

        c.drawCentredString(x_start + 300, y_start + 810, 'ΠΙΝΑΚΑΣ ΟΛΩΝ ΤΟΝ ΥΛΙΚΩΝ')

        items_data = [
            ['Εικονα', 'Κωδικος', 'Ποσοτητα', 'Κατηγορια', 'Θεση', 'Πηρα', 'Αλλο'],
            [it_424_czn10_r_afb1, '424CZN10R', (((int((panel_number_value.get())) +
                                                  int((panel_number_value2.get()))) * 2) - 2),
             'ΜΕΝΤΕΣΕΔΕΣ', '5A', '', '.................................'],
            [it_424hzn10r, '424HZN10R', (((int(panel_number_value.get()) + int(panel_number_value2.get())) * 2) - 2),
             'ΜΕΝΤΕΣΕΔΕΣ ΜΕΣΑΙΟΙ', '5A', '', '.................................'],
            [it_1085n_2510, '1085N-2510', height_value.get() + 'mm 2 ΤΕΜ', 'ΛΑΣΤΙΧΑ ΠΛΑΙΝΑ', '5A', '',
             '.................................'],
            # [it_215c_b6500, '215C-B6500', height_value.get() + 'mm 1 SET', 'ΟΔΗΓΟΣ ΜΕ ΚΟΥΡΜΠΑ', '5A'],
            [it_rscvb50z, vertical_track_code, height_value.get() + 'mm 1 SET', 'ΚΑΘΕΤΟΙ ΟΔΗΓΟΙ', '5A', '',
             '.................................'],
            [it_rschm50, horizontal_track_code, height_value.get() + 'mm 1 SET', 'ΟΡΙΖΟΝΤΙΟΙ ΟΔΗΓΟΙ', '5A', '',
             '.................................'],
            [it_870rs, '870RS', '2', 'ΣΥΝΔΕΣΜΟΣ ΟΔΗΓΩΝ', '5A', '', '.................................'],
            [it_574rn, '574RN', ((int(panel_number_value.get()) + int(panel_number_value2.get())) * 2), 'ΡΟΔΑΚΙΑ',
             '5A', '', '.................................'],
            # [it_574s, '574S', '2', 'ΡΟΔΑΚΙΑ ΠΑΝΩ', '5A'],
            [it_liftmaster, 'LIFTMASTER GRIP', '1', 'ΧΕΡΟΥΛΙ', '5A', '', '.................................'],
            [it_usa8r, 'USA8R', '1', 'ΣΤΗΡΙΓΜΑ ΑΞΟΝΑ ΔΕΞΙ', '5A', '', '.................................'],
            [it_usa8l, 'USA8L', '1', 'ΣΤΗΡΙΓΜΑ ΑΞΟΝΞΑ ΑΡΙΣΤΕΡΟ', '5A', '', '.................................'],
            [it_421k, '421K', '1 SET', 'ΚΑΤΩ ΠΙΑΣΙΜΟ', '5A', '', '.................................'],
            [it_usa_b, 'USA-B', '2', 'ΡΟΥΛΕΜΑΝ', '5A', '', '.................................'],
            [it_ff05nl12, 'FF05NL12', '2', 'ΤΥΜΠΑΝΑ', '5A', '', '.................................'],
            [it_k24_4800_ts, 'ΣΥΡΜΑΤΟΣΧΟΙΝΑ ' + lifting_cable_types_combobox.get(), str(int(height_value.get()) + 500) +
             'mm 2 ΤΕΜΑΧΙΑ', 'ΣΥΡΜΑΤΟΣΧΟΙΝΑ', '5A', '', '.................................'],
            [it_vl45_1_300gs, 'ΕΛΑΤΗΡΙΑ', '1 Η 2', 'ΕΛΑΤΗΡΙΑ', '5A', '', '.................................'],
            [it_1036, 'ΠΑΝΩ ΛΑΣΤΙΧΟ', width_value.get() + 'mm', 'ΠΑΝΩ ΛΑΣΤΙΧΟ', '5A', '',
             '.................................'],
            [it_1035_dw, 'ΚΑΤΩ ΛΑΣΤΙΧΟ', width_value.get() + 'mm', 'ΚΑΤΩ ΛΑΣΤΙΧΟ', '5A', '',
             '.................................'],
            [it_a1f, 'ΤΕΛΕΙΩΜΑΤΑ ΓΙΑ 500mm', 'L:' + (panel_number_value.get()) + ' R:' + (panel_number_value.get()),
             'ΠΛΑΙΝΑ ΑΛΟΥΜΗΝΙΑ', '5A', '', '.................................'],
            [it_a1f, 'ΤΕΛΕΙΩΜΑΤΑ ΓΙΑ 610mm', 'L:' + (panel_number_value2.get()) + ' R:' + (panel_number_value2.get()),
             'ΠΛΑΙΝΑ ΑΛΟΥΜΗΝΙΑ', '5A', '', '.................................'],
            # [it_1085rbn2035, 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', height_value.get(), 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', '5A'],
            [aluminium_profiles, 'ΠΑΝΩ & ΚΑΤΩ ΑΛΟΥΜΗΝΙΑ', '2x' + width_value.get(), 'ΠΑΝΩ & ΚΑΤΩ ΑΛΟΥΜΗΝΙΑ', '5A', '',
             '.................................'],
            [it_702gbl750, 'ΑΞΟΝΑΣ', int(width_value.get()) + 240, 'ΑΞΟΝΑΣ', '5A', '',
             '.................................'],
            [it_703st, 'ΣΥΝΔΕΣΜΟΣ ΑΞΟΝΑ', '0', 'ΣΥΝΔΕΣΜΟΣ ΑΞΟΝΑ', '5A', '', '.................................'],
            [it_629v, '629V', '0', 'ΚΛΕΙΔΩΜΑ', '5A', '', '.................................'],
            [it_italdoor_panel, 'ΠΑΝΕΛ', int(panel_number_value.get()) + int(panel_number_value2.get()), 'ΠΑΝΕΛ', '5A'],
            [moter, 'ΜΟΤΕΡ', moter_yes_or_no, 'ΜΟΤΕΡ', '5Α', '', '.................................']
        ]
        items_data = Table(items_data)
        items_data.setStyle(table_style2)
        items_data.wrapOn(c, 600, 250)
        items_data.drawOn(c, width - 585, height - 832)
    elif checkbox_panel_data == 1 and checkbox_all_data == 0:
        c.drawCentredString(x_start + 300, y_start + 805, 'ΠΙΝΑΚΑΣ ΥΛΙΚΩΝ ΠΑΝΕΛ')
        items_data = [
            ['Εικονα', 'Κωδικος', 'Ποσοτητα', 'Κατηγορια', 'Θεση', 'Πηρα', 'Αλλο'],

            [it_1036, 'ΠΑΝΩ ΛΑΣΤΙΧΟ', width_value.get() + 'mm', 'ΠΑΝΩ ΛΑΣΤΙΧΟ', '5A', '',
             '.................................'],
            [it_1035_dw, 'ΚΑΤΩ ΛΑΣΤΙΧΟ', width_value.get() + 'mm', 'ΚΑΤΩ ΛΑΣΤΙΧΟ', '5A', '',
             '.................................'],
            [it_a1f, 'ΤΕΛΕΙΩΜΑΤΑ ΓΙΑ 500mm', 'L:' + (panel_number_value.get()) + ' R:' + (panel_number_value.get()),
             'ΠΛΑΙΝΑ ΑΛΟΥΜΗΝΙΑ', '5A', '', '.................................'],
            [it_a1f, 'ΤΕΛΕΙΩΜΑΤΑ ΓΙΑ 610mm', 'L:' + (panel_number_value2.get()) + ' R:' + (panel_number_value2.get()),
             'ΠΛΑΙΝΑ ΑΛΟΥΜΗΝΙΑ', '5A', '', '.................................'],
            [aluminium_profiles, 'ΠΑΝΩ & ΚΑΤΩ ΑΛΟΥΜΗΝΙΑ', '2x' + width_value.get(), 'ΠΑΝΩ & ΚΑΤΩ ΑΛΟΥΜΗΝΙΑ', '5A', '',
             '.................................'],
            [it_italdoor_panel, 'ΠΑΝΕΛ', int(panel_number_value.get()) + int(panel_number_value2.get()), 'ΠΑΝΕΛ', '5A',
             '', '.................................'],
        ]
        items_data = Table(items_data)
        items_data.setStyle(table_style2)
        items_data.wrapOn(c, 600, 250)
        items_data.drawOn(c, width - 555, height - 265)
        print('ITEMS ONLY FOR PANEL')
    elif checkbox_rail_data == 1 and checkbox_all_data == 0:
        c.drawCentredString(x_start + 300, y_start + 805, 'ΠΙΝΑΚΑΣ ΥΛΙΚΩΝ ΟΔΗΓΩΝ')
        items_data = [
            ['Εικονα', 'Κωδικος', 'Ποσοτητα', 'Κατηγορια', 'Θεση', 'Πηρα', 'Αλλο'],
            [it_1085n_2510, '1085N-2510', height_value.get() + 'mm 2 ΤΕΜ', 'ΛΑΣΤΙΧΑ ΠΛΑΙΝΑ', '5A', '',
             '.................................'],
            # [it_215c_b6500, '215C-B6500', height_value.get() + 'mm 1 SET', 'ΟΔΗΓΟΣ ΜΕ ΚΟΥΡΜΠΑ', '5A'],
            [it_rscvb50z, vertical_track_code, height_value.get() + 'mm 1 SET', 'ΚΑΘΕΤΟΙ ΟΔΗΓΟΙ', '5A', '',
             '.................................'],
            [it_rschm50, horizontal_track_code, height_value.get() + 'mm 1 SET', 'ΟΡΙΖΟΝΤΙΟΙ ΟΔΗΓΟΙ', '5A', '',
             '.................................'],
            [it_870rs, '870RS', '2', 'ΣΥΝΔΕΣΜΟΣ ΟΔΗΓΩΝ', '5A', '', '.................................'],

            # [it_574s, '574S', '2', 'ΡΟΔΑΚΙΑ ΠΑΝΩ', '5A'],
            # [it_1085rbn2035, 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', height_value.get(), 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', '5A'],
        ]
        items_data = Table(items_data)
        items_data.setStyle(table_style2)
        items_data.wrapOn(c, 600, 250)
        items_data.drawOn(c, width - 525, height - 205)
        print('ITEMS ONLY FOR RAIL')

    elif checkbox_shaft_data == 1 and checkbox_all_data == 0:
        c.drawCentredString(x_start + 300, y_start + 805, 'ΠΙΝΑΚΑΣ ΥΛΙΚΩΝ ΑΞΟΝΑ')
        items_data = [
            ['Εικονα', 'Κωδικος', 'Ποσοτητα', 'Κατηγορια', 'Θεση', 'Πηρα', 'Αλλο'],
            [it_usa8r, 'USA8R', '1', 'ΣΤΗΡΙΓΜΑ ΑΞΟΝΑ ΔΕΞΙ', '5A', '', '.................................'],
            [it_usa8l, 'USA8L', '1', 'ΣΤΗΡΙΓΜΑ ΑΞΟΝΞΑ ΑΡΙΣΤΕΡΟ', '5A', '', '.................................'],
            [it_usa_b, 'USA-B', '2', 'ΡΟΥΛΕΜΑΝ', '5A', '', '.................................'],
            [it_ff05nl12, 'FF05NL12', '2', 'ΤΥΜΠΑΝΑ', '5A', '', '.................................'],
            [it_k24_4800_ts, 'ΣΥΡΜΑΤΟΣΧΟΙΝΑ ' + lifting_cable_types_combobox.get(), str(int(height_value.get()) + 500) +
             'mm 2 ΤΕΜΑΧΙΑ', 'ΣΥΡΜΑΤΟΣΧΟΙΝΑ', '5A', '', '.................................'],
            [it_vl45_1_300gs, 'ΕΛΑΤΗΡΙΑ', '1 Η 2', 'ΕΛΑΤΗΡΙΑ', '5A', '', '.................................'],
            # [it_1085rbn2035, 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', height_value.get(), 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', '5A'],
            [it_702gbl750, 'ΑΞΟΝΑΣ', int(width_value.get()) + 240, 'ΑΞΟΝΑΣ', '5A', '',
             '.................................'],
            [it_703st, 'ΣΥΝΔΕΣΜΟΣ ΑΞΟΝΑ', '0', 'ΣΥΝΔΕΣΜΟΣ ΑΞΟΝΑ', '5A', '', '.................................'],
        ]
        items_data = Table(items_data)
        items_data.setStyle(table_style2)
        items_data.wrapOn(c, 600, 250)
        items_data.drawOn(c, width - 560, height - 320)
        print('ITEMS ONLY FOR SHAFT')
    elif checkbox_general_data == 1 and checkbox_all_data == 0:
        c.drawCentredString(x_start + 300, y_start + 805, 'ΠΙΝΑΚΑΣ ΕΞΑΡΤΗΜΑΤΩΝ')
        items_data = [
            ['Εικονα', 'Κωδικος', 'Ποσοτητα', 'Κατηγορια', 'Θεση', 'Πηρα', 'Αλλο'],
            [it_424_czn10_r_afb1, '424CZN10R', (((int((panel_number_value.get())) +
                                                  int((panel_number_value2.get()))) * 2) - 2),
             'ΜΕΝΤΕΣΕΔΕΣ', '5A', '', '.................................'],
            [it_424hzn10r, '424HZN10R', (((int(panel_number_value.get()) + int(panel_number_value2.get())) * 2) - 2),
             'ΜΕΝΤΕΣΕΔΕΣ ΜΕΣΑΙΟΙ', '5A', '', '.................................'],

            [it_574rn, '574RN', ((int(panel_number_value.get()) + int(panel_number_value2.get())) * 2), 'ΡΟΔΑΚΙΑ',
             '5A', '', '.................................'],
            # [it_574s, '574S', '2', 'ΡΟΔΑΚΙΑ ΠΑΝΩ', '5A'],
            [it_liftmaster, 'LIFTMASTER GRIP', '1', 'ΧΕΡΟΥΛΙ', '5A', '', '.................................'],

            [it_421k, '421K', '1 SET', 'ΚΑΤΩ ΠΙΑΣΙΜΟ', '5A', '', '.................................'],

            # [it_1085rbn2035, 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', height_value.get(), 'ΠΛΑΙΝΑ ΛΑΣΤΙΧΑ', '5A'],
            [it_629v, '629V', '0', 'ΚΛΕΙΔΩΜΑ', '5A', '', '.................................'],
            [moter, 'ΜΟΤΕΡ', moter_yes_or_no, 'ΜΟΤΕΡ', '5Α', '', '.................................']
        ]
        items_data = Table(items_data)
        items_data.setStyle(table_style2)
        items_data.wrapOn(c, 600, 250)
        items_data.drawOn(c, width - 520, height - 285)
        print('ONLY GENERAL ITEMS')

    # items_data = Table(items_data)
    # items_data.setStyle(table_style2)
    # items_data.wrapOn(c, 600, 250)
    # items_data.drawOn(c, width - 585, height - 805)
    # TELOS 27/12/2022 TABLE OF CONSUMABLES WITH IMAGE
    c.save()

    # Open the created PDF file with the default selected program
    os.startfile(file_path + folder_name + '\\' + name_value.get() + '_' + f"{date:%d-%m-%Y}" + '.pdf')

    return


# Συμπλήρωση τιμής βάρους πάνελ
def on_field_change(index, value, op):
    # print("combobox updated to ", panel_types_combobox.get())
    if panel_types_combobox.current() == 0:
        panel_weight_value.delete(0, END)
        panel_weight_value.insert(0, '10.5')
    else:
        panel_weight_value.delete(0, END)
        panel_weight_value.insert(0, '0.0')
    # Με αυτό κλειδώνουμε την τιμή να μην μπορεί να την αλλάξει ο χρήστης
    # panel_weight_value.config(state=DISABLED)


# Συνολικό Βάρος Πανελ Υπολογισμός
def on_field_change_total_value():
    try:

        if panel_types_combobox.current() == 0:
            int(width_value.get())
            int(height_value.get())
            print('Υπολογισμός Συνολικού βάρους τρέχον width value ', width_value.get())
            print('Υπολογισμός Συνολικού βάρους τρέχον height value ', height_value.get())
            total_panel_weight = 10.5 * ((float(width_value.get()) + 40) / 1000) * (float(height_value.get()) / 1000)
            total_panel_weight_formated = format(total_panel_weight, '.2f')
            panel_weight_total_value.delete(0, END)
            panel_weight_total_value.insert(0, str(total_panel_weight_formated))

            # height_cut_value.get()
            # height_cut_value.delete(0, END)
        # height_cut_value_on_def = str(((int(panel_number_value.get()) * 500 + int(panel_number_value2.get()) * 610)
        #                                - int(height_value.get())))
        # height_cut_value.insert(END, height_cut_value_on_def)
        # print('Η ΚΟΠΗ ΠΟΥ ΘΑ ΓΙΝΕΙ ΣΤΟ ΥΨΟΣ ΕΙΝΑΙ:' + str(height_cut_value.get()))
        # koph_panw_panel.get()
        # koph_panw_panel.delete(0, END)
        # if int(panel_number_value2.get()) > 0:
        #     koph_panw_panel.insert(END, str(610 - int(height_cut_value_on_def)))
        #     print('ΤΙΜΗ ΚΟΠΗ ΠΑΝΩ ΠΑΝΕΛ (610-koph_panw_panel):' + koph_panw_panel.get())
        # elif int(panel_number_value.get()) > 0:
        #     koph_panw_panel.insert(END, str(500 - int(height_cut_value_on_def)))
        #     print(koph_panw_panel.get())
    except ValueError:
        return


# Υπολογισμός Πανελ που χρειαζόμαστε
def on_field_change_height():
    koph_panw_panel = Entry(root)
    koph_panw_panel.insert(END, '0')
    try:
        int(height_value.get())
        if int(height_value.get()) != 0:
            if int(height_value.get()) == 1500:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '3')
                panel_number_value2.delete(0, END)
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2610 - int(height_value.get())))
            elif 1501 <= int(height_value.get()) <= 1610:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '2')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '1')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2610 - int(height_value.get())))
            elif 1611 <= int(height_value.get()) <= 1720:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '1')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '2')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(1720 - int(height_value.get())))
            elif 1721 <= int(height_value.get()) <= 1830:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '0')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '3')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(1830 - int(height_value.get())))
            elif 1831 <= int(height_value.get()) <= 2000:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '4')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '0')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2000 - int(height_value.get())))
            elif 2001 <= int(height_value.get()) <= 2110:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '3')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '1')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2110 - int(height_value.get())))
            elif 2111 <= int(height_value.get()) <= 2220:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '2')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '2')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2220 - int(height_value.get())))
            elif 2221 <= int(height_value.get()) <= 2330:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '1')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '3')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2330 - int(height_value.get())))
            elif 2331 <= int(height_value.get()) <= 2440:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '0')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '4')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2440 - int(height_value.get())))
            elif 2441 <= int(height_value.get()) <= 2500:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '5')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '0')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2500 - int(height_value.get())))
            elif 2501 <= int(height_value.get()) <= 2610:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '4')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '1')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2610 - int(height_value.get())))
            elif 2611 <= int(height_value.get()) <= 2720:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '3')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '2')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2720 - int(height_value.get())))
            elif 2721 <= int(height_value.get()) <= 2830:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '2')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '3')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2830 - int(height_value.get())))
            elif 2831 <= int(height_value.get()) <= 2940:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '1')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '4')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(2940 - int(height_value.get())))
            elif 2941 <= int(height_value.get()) <= 3000:
                panel_number_value.delete(0, END)
                panel_number_value.insert(END, '0')
                panel_number_value2.delete(0, END)
                panel_number_value2.insert(END, '5')
                koph_panw_panel.delete(0, END)
                koph_panw_panel.insert(END, str(3000 - int(height_value.get())))
    except ValueError:
        return koph_panw_panel.get()

    # # panel_number_610mm=0
    # try:
    #     int(height_value.get())
    #     if height_value.get() != 0:
    #         if int(height_value.get()) % 500 == 0:
    #             panel_number_500 = int(height_value.get()) / 500
    #             print("ο αριθμός των πάνελ 500mm που θα χρησιμοποιηθούν είναι:", panel_number_500)
    #             panel_number_value2.delete(0, END)
    #             panel_number_value2.insert(END, '0')
    #             panel_number_value.delete(0, END)
    #             panel_number_value.insert(END, str(panel_number_500))
    #         elif int(height_value.get()) % 610 == 0:
    #             panel_number_610 = int(height_value.get()) / 610
    #             print("ο αριθμός των πάνελ 610mm που θα χρησιμοποιηθούν είναι:", panel_number_610)
    #             panel_number_value.delete(0, END)
    #             panel_number_value.insert(END, '0')
    #             panel_number_value2.delete(0, END)
    #             panel_number_value2.insert(END, str(panel_number_610))
    #         elif int(height_value.get()) % 500 <= 100:
    #             panel_number_500 = (int(height_value.get()) // 500) - 1
    #             panel_number_610 = 1
    #             print("ο αριθμός των 500mm πάνελ που θα χρησιμοποιηθούν είναι:", panel_number_500,
    #                   ' και ο αριθμός των 610μμ πάνελ που θα χρησιμοποιηθούν είναι:1')
    #             panel_number_value.delete(0, END)
    #             panel_number_value2.delete(0, END)
    #             panel_number_value.insert(END, str(panel_number_500))
    #             panel_number_value2.insert(END, str(panel_number_610))
    #         elif int(height_value.get()) % 500 > 110:
    #             panel_number_500 = (int(height_value.get()) // 500) + 1
    #             print("ο αριθμός των πάνελ 610mm που θα χρησιμοποιηθούν είναι:", panel_number_500)
    #             panel_number_value.delete(0, END)
    #             panel_number_value2.delete(0, END)
    #             panel_number_value2.insert(END, '0')
    #             panel_number_value.insert(END, str(panel_number_500))
    # except ValueError:
    #     return


# Έλεγχος τιμής ύψους
def number():
    try:
        int(height_value.get())
        if 3000 >= int(height_value.get()) != 0:
            answer.config(text="Σωστή διάσταση!")
        else:
            answer.config(text='Λάθος Τιμή')
        # print(height_value.get())
    except ValueError:
        answer.config(text='Λάθος Τιμή')


# έλεγχος τιμής πλάτους
def number2():
    try:
        int(width_value.get())
        if 5000 >= int(width_value.get()) != 0:
            answer2.config(text="Σωστή Διάσταση")
        else:
            answer2.config(text='Λάθος Τιμή')
    except ValueError:
        answer2.config(text='Λάθος Τιμή')


# Υπολογισμός Κοπής Ύψους
# def height_cut_calculation():
#     height_cut_value = ((int(panel_number_value.get()) * 500 + int(panel_number_value2.get()) * 610)
#                         - int(height_value.get()))
#     print('Η ΚΟΠΗ ΠΟΥ ΘΑ ΓΙΝΕΙ ΣΤΟ ΥΨΟΣ ΕΙΝΑΙ:' + str(height_cut_value))


logo_label = Label(general_tab, image=new_resized_imagelogo)
logo_label.grid(column=0, row=0)
label = ttk.Label(text="Υπολογισμός Υλικών Κατασκευής Πόρτας")

# Στοιχεία Πελάτη


blank1 = Label(client_tab, text='', font=('Times New Roman', 12))
blank1.grid(row=1, column=1, sticky='e')
blank3 = Label(client_tab, text='', font=('Times New Roman', 12))
blank3.grid(row=3, column=1, sticky='e')
blank5 = Label(client_tab, text='', font=('Times New Roman', 12))
blank5.grid(row=5, column=1, sticky='e')
blank7 = Label(client_tab, text='', font=('Times New Roman', 12))
blank7.grid(row=7, column=1, sticky='e')
blank9 = Label(client_tab, text='', font=('Times New Roman', 12))
blank9.grid(row=9, column=1, sticky='e')
# name_value = Entry(client_tab, width=30)
# name_value.grid(row=2, column=2, sticky='w')
# Ημερομηνία
date = dt.datetime.now()
date_label = Label(client_tab, text='Ημερομηνία:', font=('Times New Roman', 12))
date_label.grid(row=0, column=0, sticky='e')
date_value = Entry(client_tab, width=10)
date_value.grid(row=0, column=1, sticky='w')
date_value.insert(END, f"{date:%d/%m/%Y}")

# Ονοματεπώνυμο
name_label = Label(client_tab, text='Πελατης:', font=('Times New Roman', 12))
name_label.grid(row=2, column=1, sticky='e')
name_value = Entry(client_tab, width=30)
name_value.grid(row=2, column=2, sticky='w')
# Εταιρεια
company_label = Label(client_tab, text='Εταιρεια:', font=('Times New Roman', 12))
company_label.grid(row=4, column=1, sticky='e')
company_value = Entry(client_tab, width=30)
company_value.grid(row=4, column=2, sticky='w')
# ΑΦΜ
afm_label = Label(client_tab, text='Α.Φ.Μ.:', font=('Times New Roman', 12))
afm_label.grid(row=6, column=1, sticky='e')
afm_value = Entry(client_tab, width=30)
afm_value.grid(row=6, column=2, sticky='w')
# ΔΟΥ
doy_label = Label(client_tab, text='ΔΟΥ:', font=('Times New Roman', 12))
doy_label.grid(row=8, column=3, sticky='e')
doy_value = Entry(client_tab, width=30)
doy_value.grid(row=8, column=4, sticky='w')
# EMAIL
email_label = Label(client_tab, text='Email:', font=('Times New Roman', 12))
email_label.grid(row=6, column=3, sticky='e')
email_value = Entry(client_tab, width=30)
email_value.grid(row=6, column=4, sticky='w')
# Phone
phone_label = Label(client_tab, text='       Τηλεφωνο:', font=('Times New Roman', 12))
phone_label.grid(row=2, column=3, sticky='e')
phone_value = Entry(client_tab, width=30)
phone_value.grid(row=2, column=4, sticky='w', )
# Σταθερο Τηλεφωνο
home_phone_label = Label(client_tab, text='Σταθερο:', font=('Times New Roman', 12))
home_phone_label.grid(row=4, column=3, sticky='e')
home_phone_value = Entry(client_tab, width=30)
home_phone_value.grid(row=4, column=4, sticky='w')
# # ΔΙΕΥΘΥΝΣΗ
# ΟΔΟΣ
address_label = Label(client_tab, text='Οδος:', font=('Times New Roman', 12))
address_label.grid(row=8, column=1, sticky='e')
address_value = Entry(client_tab, width=30)
address_value.grid(row=8, column=2, sticky='w')
# ΠΟΛΗ
city_label = Label(client_tab, text='Πολη:', font=('Times New Roman', 12))
city_label.grid(row=10, column=3, sticky='e')
city_value = Entry(client_tab, width=30)
city_value.insert(END, 'Θεσσαλονικη')
city_value.grid(row=10, column=4, sticky='w')
# TK
zip_code_label = Label(client_tab, text='Τ.Κ.:', font=('Times New Roman', 12))
zip_code_label.grid(row=10, column=1, sticky='e')
zip_code_value = Entry(client_tab, width=30)
zip_code_value.grid(row=10, column=2, sticky='w')
# ΧΩΡΑ
country_label = Label(client_tab, text='Χωρα:', font=('Times New Roman', 12))
country_label.grid(row=12, column=3, sticky='e')
country_value = Entry(client_tab, width=30)
country_value.insert(END, 'Ελλαδα')
country_value.grid(row=12, column=4, sticky='w')

# ΤΕΛΟΣ Στοιχεία Πελάτη

# Διαχείριση ύψους με έλεγχο τιμής από την def number
dimensions_height = Label(general_tab, text='Υψος (mm):', font=('Times New Roman', 12))
dimensions_height.grid(column=0, row=3, sticky='e')
check_button_height = Button(general_tab, text="Έλεγχος Τιμής", command=number)
check_button_height.grid(column=2, row=3, sticky='w')
height_value = Entry(general_tab, width=5)
height_value.grid(column=1, row=3, sticky='w')
height_value.insert(END, '0')
answer = Label(general_tab, text='')
answer.grid(column=3, row=3, sticky='w')

# Διαχείριση πλάτους με έλεγχο τιμής από την def number2
dimensions_width = Label(general_tab, text="Πλατος (mm):", font=('Times New Roman', 12))
dimensions_width.grid(column=0, row=2, sticky='e')
check_button_width = Button(general_tab, text="Έλεγχος Τιμής", command=number2)
check_button_width.grid(column=2, row=2, sticky='w')
width_value = Entry(general_tab, width=5)
width_value.grid(column=1, row=2, sticky='w')
width_value.insert(END, '0')
answer2 = Label(general_tab, text='')
answer2.grid(column=3, row=2, sticky='w')

# Πάνελ που θα χρησιμοποιειθεί
panel_type_label = Label(general_tab, text='Τυπος Πανελ:', font=('Times New Roman', 12))
panel_types = ['Italpannelli', 'αλλο']
var.trace('w', on_field_change)
panel_types_combobox = ttk.Combobox(general_tab, textvar=var, values=panel_types, width=10)
panel_type_label.grid(column=0, row=4, sticky='e')
panel_types_combobox.grid(row=4, column=1, sticky='w')
panel_types_combobox['state'] = 'readonly'
panel_types_combobox.set('')
panel_weight_label = Label(general_tab, text="Βαρος Πανελ kg/τμ:", font=('Times New Roman', 12))
panel_weight_label.grid(column=0, row=5, sticky='e')
panel_weight_value = Entry(general_tab, width=5)
panel_weight_value.grid(column=1, row=5, sticky='w')

# Συνολικό Βάρος Πάνελ
panel_weight_total = Label(general_tab, text='Συνολικο Βαρος Πανελ (kg):', font=('Times New Roman', 12))
panel_weight_total.grid(column=0, row=6, sticky='e')
panel_weight_total_value = Entry(general_tab, width=10)
panel_weight_total_value.grid(column=1, row=6, sticky='w')
panel_weight_button = Button(general_tab, text='Βαρος', command=on_field_change_total_value)
panel_weight_button.grid(row=7, column=2, sticky='w')

# Λειτουργία Πόρτας (Χειροκίνητη ή Μοτέρ)
operation = Label(general_tab, text='Λειτουργια:', font=('Times New Roman', 12))
operation.grid(row=2, column=4, sticky='e')
operation_types = ['Χειροκινητη', 'Μοτερ', 'Αλυσιδα']
operation_combobox = ttk.Combobox(general_tab, values=operation_types, width=12)
operation_combobox['state'] = 'readonly'
operation_combobox.grid(row=2, column=5, sticky='w')

# Μοντέλο Μοτέρ
moter_model = Label(general_tab, text='Μοντελο Μοτερ:', font=('Times New Roman', 12))
moter_model.grid(row=3, column=4, sticky='e')
moter_model_values = ['KIT C40/1000', 'SET C40/3301', 'SET C40/3301/HR', 'SET C40/4001', 'SET C40/5001']
moter_model_combobox = ttk.Combobox(general_tab, values=moter_model_values, width=16)
moter_model_combobox['state'] = 'readonly'
moter_model_combobox.grid(row=3, column=5, sticky='w')

# Τύπος Οδηγών
lift_system = Label(general_tab, text='Τυπος Πορτας:', font=('Times New Roman', 12))
lift_system.grid(row=4, column=4, sticky='e')
lift_system_types = ['SL', 'LH', 'VL', 'HL', 'HL-P', 'SLP']
lift_system_types_combobox = ttk.Combobox(general_tab, values=lift_system_types, width=6)
lift_system_types_combobox['state'] = 'readonly'
lift_system_types_combobox.grid(row=4, column=5, sticky='w')

# Αριθμός πάνελ που θα χρησιμοποιηθούν σύμφωνα με το ύψος

# 500mm italpannelli
# var.trace('w', on_field_change_height)
panel_number_500mm = Label(general_tab, text='Αριθμος Πανελ των 500mm:', font=('Times New Roman', 12))
panel_number_500mm.grid(row=7, column=0, sticky='e')
panel_number_500mm_button = Button(general_tab, text='Υπολογισε Πανελ', command=on_field_change_height)
panel_number_500mm_button.grid(row=6, column=2, sticky='w')
panel_number_value = Entry(general_tab, width=5)
panel_number_value.grid(column=1, row=7, sticky='w')
panel_number_value.insert(END, '0')
# 610mm italpannelli
panel_number_610mm = Label(general_tab, text='Αριθμος Πανελ των 610mm:', font=('Times New Roman', 12))
panel_number_610mm.grid(row=8, column=0, sticky='e')
panel_number_value2 = Entry(general_tab, width=5)
panel_number_value2.grid(column=1, row=8, sticky='w')
panel_number_value2.insert(END, '0')

# Χρώμα Πόρτας
color_label = Label(general_tab, text='Κωδικος Χρωματος:', font=('Times New Roman', 12))
color_label.grid(row=9, column=0, sticky='e')
color_value = Entry(general_tab, width=8)
color_value.grid(row=9, column=1, sticky='w')
color_value.insert(END, 'RAL ')

# ΣΥΡΜΑΤΟΣΧΟΙΝΑ
lifting_cable = Label(general_tab, text="Συρματοσχοινα:", font=('Times New Roman', 12))
lifting_cable.grid(row=7, column=4, sticky='e')
lifting_cable_types = ['3mm', '4mm', '5mm', '6mm']
lifting_cable_types_combobox = ttk.Combobox(general_tab, values=lifting_cable_types, width=5)
lifting_cable_types_combobox['state'] = 'readonly'
lifting_cable_types_combobox.grid(column=5, row=7, sticky='w')

# DECTION
dection = Label(general_tab, text="Dection:", font=('Times New Roman', 12))
dection.grid(row=8, column=4, sticky='e')
dection_types = ['Ναι', 'Οχι']
dection_types_combobox = ttk.Combobox(general_tab, values=dection_types, width=5)
dection_types_combobox['state'] = 'readonly'
dection_types_combobox.grid(column=5, row=8, sticky='w')


def on_checkbox_click(var):
    print("Checkbox clicked, value is now", var.get())


checkbox_var_all = IntVar(value=1)  # set the default value of print all (checkbox_all to 1)
checkbox_var_shaft = IntVar(value=0)  # set the default value of print shaft (checkbox_all to 0)
checkbox_var_panel = IntVar(value=0)
checkbox_var_rail = IntVar(value=0)
checkbox_var_general = IntVar(value=0)

checkbox_all = Checkbutton(general_tab, text="Εκτύπωση ΟΛΩΝ", variable=checkbox_var_all,
                           command=lambda: on_checkbox_click(checkbox_var_all))
checkbox_shaft = Checkbutton(general_tab, text="Εκτύπωση MONO ΑΞΟΝΑ", variable=checkbox_var_shaft,
                             command=lambda: on_checkbox_click(checkbox_var_shaft))
checkbox_panel = Checkbutton(general_tab, text="Εκτύπωση MONO ΠΑΝΕΛ", variable=checkbox_var_panel,
                             command=lambda: on_checkbox_click(checkbox_var_panel))
checkbox_rail = Checkbutton(general_tab, text="Εκτύπωση MONO ΟΔΗΓΟΙ", variable=checkbox_var_rail,
                            command=lambda: on_checkbox_click(checkbox_var_rail))
checkbox_general = Checkbutton(general_tab, text="Εκτύπωση MONO ΕΞΑΡΤΗΜΑΤΑ", variable=checkbox_var_general,
                               command=lambda: on_checkbox_click(checkbox_var_general))

checkbox_all.grid(column=5, row=9, sticky='w')
checkbox_shaft.grid(column=5, row=10, sticky='w')
checkbox_panel.grid(column=5, row=11, sticky='w')
checkbox_rail.grid(column=5, row=12, sticky='w')
checkbox_general.grid(column=5, row=13, sticky='w')

# Συνδεσμος Αξονα
shaft_coupler = Label(general_tab, text='Συνδεσμος Αξονα:', font=('Times New Roman', 12))
shaft_coupler.grid(column=4, row=6, sticky='e')
shaft_coupler_values = ['Ναι', 'Οχι']
shaft_coupler_combobox = ttk.Combobox(general_tab, values=shaft_coupler_values, width=5)
shaft_coupler_combobox.current(1)
shaft_coupler_combobox['state'] = 'readonly'
shaft_coupler_combobox.grid(column=5, row=6, sticky='w')

# Κύκλοι ζωής ελατηρίων
# spring_life = Label(general_tab, text='Κύκλοι ζωής ελατηρίων:', font=('Times New Roman', 12))
# spring_life.grid(row=7, column=4, sticky='e')
# spring_life_values = ['12.500', '15.000', '25.000', '35.000', '50.000', '100.000']
# spring_life_combobox = ttk.Combobox(general_tab, values=spring_life_values, width=8)
# spring_life_combobox['state'] = 'readonly'
# spring_life_combobox.grid(row=7, column=5, sticky='w')

# Εσωτερική Διάμετρός Ελατηριου
# spring_diameter = Label(general_tab, text='Εσωτερική Διάμετρος Ελατηρίου:', font=('Times New Roman', 12))
# spring_diameter.grid(row=8, column=4, sticky='e')
# spring_diameter_values = ['50,8mm', '66,68mm', '95,25mm', '152,4mm']
# spring_diameter_combobox = ttk.Combobox(general_tab, values=spring_diameter_values, width=9)
# spring_diameter_combobox['state'] = 'readonly'
# spring_diameter_combobox.grid(row=8, column=5, sticky='w')

# Τυπος Αξονα
shaft_type = Label(general_tab, text='Τύπος Άξονα:', font=('Times New Roman', 12))
shaft_type.grid(row=5, column=4, sticky='e')
shaft_type_values = ['3mm Κουφιος με ασφαλεια', '2mm Κουφιος']
shaft_type_values_combobox = ttk.Combobox(general_tab, values=shaft_type_values, width=25)
shaft_type_values_combobox['state'] = 'readonly'
shaft_type_values_combobox.grid(column=5, row=5, sticky='w')

# Ενίσχυση κάτω Μέρους
bottom_reinforcement = Label(general_tab, text='Ενίσχυση Κάτω Πάνελ:', font=('Times New Roman', 12))
bottom_reinforcement.grid(row=10, column=0, sticky='e')
bottom_reinforcement_types = ['Ναι', 'Οχι']
bottom_reinforcement_combobox = ttk.Combobox(general_tab, values=bottom_reinforcement_types, width=5)
bottom_reinforcement_combobox['state'] = 'readonly'
bottom_reinforcement_combobox.grid(column=1, row=10, sticky='w')
height_cut_value = Entry(root)
height_cut_value.insert(END, '0')
# koph_panw_panel = Entry(root)
# koph_panw_panel.insert(END, '0')

# Κουμπί για να κάνει extract τα user input σε pdf
sto = Style()
sto.configure('W.TButton', font=('Arial', 11), foreground='Red', activebackground='#655')
pdf_button = ttk.Button(general_tab, text='Δημιουργησε PDF', style='W.TButton', command=lambda: gen_pdf(), width=20)
pdf_button.grid(row=15, column=2, sticky='s', pady=20, padx=20)

# go to second page button
# i=8
for i in range(8, 17, 1):
    blank = Label(client_tab, text='')
    blank.grid(row=i, column=0, sticky='e')

next_page_button = ttk.Button(client_tab, text='Επόμενο Βήμα', style='W.TButton', command=lambda: my_notebook.
                              select(general_tab), width=15)
next_page_button.grid(row=14, column=0, sticky='w', padx=10)

# go to first page button
next_page_button = ttk.Button(general_tab, text='Προηγούμενο Βήμα', style='W.TButton', command=lambda: my_notebook.
                              select(client_tab), width=18)
next_page_button.grid(row=15, column=0, sticky='w', padx=10, pady=10)

tab = ttk.Frame(my_notebook)
my_notebook.add(tab, text="ΦΑΚΕΛΟΣ")

# Create a listbox to show the PDF files
listbox = Listbox(tab, selectmode='single')
listbox.pack(side='left', fill='both', expand=True)

# Populate the listbox with the PDF files in the directory
directory = "C:\\python\\NADI\\grammh_paragwghs v2.0\\"
excluded_folder_names = ["venv", "RAL Garage Door", "files", "dist", "build", ".idea"]
for file in os.listdir(directory):
    if os.path.isdir(file) == TRUE and file not in excluded_folder_names:
        # if file.endswith('.pdf'):
        listbox.insert('end', file)


# Create a button to open the selected file
def open_file():
    selected_index = listbox.curselection()
    if len(selected_index) > 0:
        file_to_open = listbox.get(selected_index[0])
        # Open the file here using your preferred method
        os.startfile(file_to_open)
        pass


open_button = ttk.Button(tab, text='Open', command=open_file)
open_button.pack(side='right')

# Pack the TabControl and start the main loop
# tab.pack(expand=1, fill='both')

# def show_folder_contents(folder):
#     # Clear the frame
#     for widget in tab3.winfo_children():
#         widget.destroy()
#
#     # Get a list of the files and directories in the folder
#     contents = os.listdir(folder)
#
#     # Add a label for each file and directory
#     for item in contents:
#         if item.endswith('.pdf'):
#             label = ttk.Label(tab3, text=item)
#             label.pack()
#
#
# #
# # # Create an entry widget and a button to specify the folder
# button_show = ttk.Button(tab3, text="Show Contents", command=lambda: show_folder_contents(file_path))
# button_show.pack()

root.mainloop()
